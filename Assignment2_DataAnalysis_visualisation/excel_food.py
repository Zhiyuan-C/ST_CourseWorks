from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
import sqlite3

# Query database and return the data
def get_data()->[()]:
  conn = sqlite3.connect('database.db')
  data = []
  with conn:
    curs = conn.cursor()
    sql = """SELECT violation_code, violation_description,count(violation_code)
            FROM violations
            GROUP BY violation_code
            ORDER BY violation_code"""
    for row in curs.execute(sql):
      data.append(row)
  conn.close()
  return data

# Create workbook and write data into the workbook
def create_workbook():
  wb = Workbook()
  ws = wb.active
  ws.title = 'Violations Types'
  ws = wb['Violations Types']
  # write the header
  ws['A1'] = "violation_code"
  ws['B1'] = "violation_description"
  ws['C1'] = "count"
  rows = get_data()
  row_num = 2
  col_num = 1

  # write data into worksheet
  for row in rows:
    for val in row:
      ws.cell(row=row_num, column=col_num).value = val
      col_num += 1
    row_num += 1
    col_num = 1

  # Adjust column width
  for col in ws.columns:
    max_length = 0
    column = col[0].column # Get the column name
    for cell in col:
      try: # Avoid error on empty cells
        if len(str(cell.value)) > max_length:
          max_length = len(str(cell.value))
      except:
        pass
    ws.column_dimensions[column].width = max_length + 1
  
  # save the file
  wb.save("ViolationTypes.xlsx")


if __name__ == '__main__':
	create_workbook()