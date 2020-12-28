import sqlite3
from openpyxl import load_workbook

# Iterate the data from database, return a list of tuples
def iterate_data(ws, row_max:int, col_max:int)->[()]:
  # Create empty list to store data
  data = []
  # Iterate through the list
  for row in ws.iter_rows(min_row = 2, max_col = col_max, max_row = row_max):
    # row_data list to store cell value in each row
    row_data = [] 
    for cell in row:
      row_data.append(cell.value)
    # update data list
    data.append(tuple(row_data))
  # return final data
  return data

# retrieve data from inspections excel file
def data_inspections()->[()]:
  wb = load_workbook('inspections.xlsx')
  ws = wb['inspections']
  row_max = ws.max_row # determine the max row
  col_max = ws.max_column # determine the max column
  return iterate_data(ws, row_max, col_max)

# retrieve data from violations excel file
def data_violations()->[()]:
  wb = load_workbook('violations.xlsx')
  ws = wb['violations']
  row_max = ws.max_row # determine the max row
  col_max = ws.max_column # determine the max column
  return iterate_data(ws, row_max, col_max)

# Create database and insert data into database
def create_database():
  print("Creating database")
  conn = sqlite3.connect('database.db')
  with conn:
    curs = conn.cursor()
    # Drop table if exist
    curs.execute("DROP TABLE IF EXISTS inspections")
    curs.execute("DROP TABLE IF EXISTS violations")

    # create table inspections
    curs.execute("""CREATE TABLE inspections(
                      activity_date DATE,
                      employee_id VARCHAR(10),
                      facility_address TEXT,
                      facility_city TEXT,
                      facility_id VARCHAR(10),
                      facility_name TEXT,
                      facility_state CHAR(5),
                      facility_zip VARCHAR(15),
                      grade CHAR(1),
                      owner_id VARCHAR(10),
                      owner_name TEXT,
                      pe_description TEXT,
                      program_element_pe INT,
                      program_name TEXT,
                      program_status VARCHAR(10),
                      record_id VARCHAR(10),
                      score INT,
                      serial_number VARCHAR(10),
                      service_code INT,
                      service_description TEXT
                    )""")
    # create table violations
    curs.execute("""CREATE TABLE violations(
                      points INT,
                      serial_number VARCHAR(10),
                      violation_code CHAR(5),
                      violation_description TEXT,
                      violation_status TEXT
                    )""")                
    # insert values
    print("Retrieving data from Excel")
    inspections = data_inspections()
    violations = data_violations()
    print("Inserting data to the database")
    # insert values into inspections
    curs.executemany("INSERT INTO inspections VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", inspections)
    # insert values into violations
    curs.executemany("INSERT INTO violations VALUES (?,?,?,?,?)", violations)
    # ------  END with conn --------
  conn.close()
  print("Successfully created database")

if __name__ == '__main__':
  create_database()
  

